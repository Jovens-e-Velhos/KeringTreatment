#!/usr/bin/env python3
"""
KERING IMPORT OPERATIONS — Gerador Standalone de Dashboard
=============================================================
Le a planilha unificada e gera um arquivo .html completo e
autossuficiente com o dashboard executivo (KPIs, pipeline,
graficos de tendencia, SLA, NF->POD, modal de delayed shipments).

USO:
    python gerar_dashboard.py CAMINHO_DA_PLANILHA.xlsx [--out SAIDA.html]

EXEMPLO:
    python gerar_dashboard.py "FUP_Kering_-_Unificado_new.xlsx"
    python gerar_dashboard.py dados.xlsx --out dashboard_kering.html

DEPENDENCIAS:
    pip install openpyxl

NAO REQUER NENHUM OUTRO ARQUIVO. NAO REQUER CONEXAO COM IA.
Todo o tratamento de dados, calculo de metricas e geracao do
HTML/CSS/JS estao embutidos neste unico arquivo .py.
"""

import argparse
import datetime
import json
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl

# ══════════════════════════════════════════════════════════════════════════
# CONFIGURACAO
# ══════════════════════════════════════════════════════════════════════════

ABA_PLANILHA = "Base Unificada"   # nome da aba a ser lida (cai para a aba ativa se nao existir)
DATA_CORTE   = datetime.datetime(2026, 4, 1)   # Invoice Receipt >= esta data

# Status considerados validos (exatamente estes 6)
STATUS_VALIDOS = {
    "WAITING ID REGISTER",
    "WAITING NF",
    "WAITING DELIVERY SCHEDULE",
    "DELIVERY SCHEDULED",
    "WAITING CARGO ATTENDANCE",
    "WAITING CUSTOMS CLEARANCE",
}
# EXCLUIDOS propositalmente: WAITING ARRIVAL, WAITING IBAMA, e qualquer outro

# Status usado para a leitura adicional de processos ja entregues
STATUS_DELIVERED = {"DELIVERED"}


# Status considerados validos para a etapa 'STATUS' no Dashboard (exatamente estes 9)
STATUS_VALIDOS_STS = {
    "WAITING ID REGISTER",
    "WAITING NF",
    "WAITING DELIVERY SCHEDULE",
    "DELIVERY SCHEDULED",
    "WAITING CARGO ATTENDANCE",
    "WAITING CUSTOMS CLEARANCE",
    "WAITING ARRIVAL",
    "WAITING GL",
    "WAITING PRE ALERT",
}

# Mapeamento de colunas da planilha (0-based)
# A=0 SHIPMENT  B=1 INVOICE  C=2 BOXES  D=3 ITEMS  E=4 STATUS
# F=5 STORE LOCATION  G=6 CHANNEL  H=7 SLA STATUS
# I=8 INVOICE RECEIPT  J=9 ETA/ATA
# K=10 lt_inv_gl  L=11 lt_arr_id  M=12 lt_id_cc
# N=13 lt_cc_nf   O=14 lt_nf_pod  P=15 CLIENTE
COL = {
    "shipment": 0, "invoice": 1, "boxes": 2, "items": 3, "status": 4,
    "loc": 5, "channel": 6, "status_lead": 7, "inv_receipt": 8, "eta": 10,
    "lt_ig": 16, "lt_ai": 18, "lt_ic": 19, "lt_cn": 21, "lt_np": 22,
    "brand": 23
}

BRANDS_ESPERADAS = ["BALENCIAGA", "BOTTEGA", "YSL", "GUCCI"]


# ══════════════════════════════════════════════════════════════════════════
# HELPERS DE PARSING
# ══════════════════════════════════════════════════════════════════════════

def parse_date(v):
    """Converte valor da celula para datetime. Aceita datetime nativo do
    Excel ou string em formato DD/MM/YYYY, YYYY-MM-DD ou MM/DD/YYYY."""
    if isinstance(v, datetime.datetime):
        return v
    if isinstance(v, str):
        v = v.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.datetime.strptime(v, fmt)
            except ValueError:
                continue
    return None


def val_lt(x):
    """Retorna o lead time (em dias) se for numerico e estiver entre 0 e
    120, caso contrario retorna None (trata erros e celulas vazias)."""
    if isinstance(x, (int, float)) and 0 <= x <= 120:
      return round(float(x), 2)
    return None


def clean_str(v):
    """Retorna a celula como string limpa (trim), ou string vazia se None."""
    return str(v).strip() if v is not None else ""


# ══════════════════════════════════════════════════════════════════════════
# EXTRACAO DA PLANILHA
# ══════════════════════════════════════════════════════════════════════════

def extrair_rows(caminho_planilha, status_filtro=None, silencioso=False):
    """Le a planilha e retorna uma lista de linhas no formato:
    [brand, ship, boxes, items, status, loc, channel, status_lead,
     inv_receipt(str YYYY-MM-DD), eta(str|None), lt_ig, lt_ai, lt_ic,
     lt_cn, lt_np, invoice]
    Aplica os filtros de data e de status.

    status_filtro: conjunto de status aceitos. Se None (padrao), usa
        STATUS_VALIDOS (comportamento original, linhas "em aberto").
        Para capturar entregas, passar {"DELIVERED"}.
    silencioso: se True, suprime os prints de progresso/depuracao
        (usado na segunda leitura, para os dados DELIVERED)."""
    if status_filtro is None:
        status_filtro = STATUS_VALIDOS

    contador = 1

    if not silencioso:
        print(f"Lendo planilha: {caminho_planilha}")
    wb = openpyxl.load_workbook(caminho_planilha, read_only=True, data_only=True)

    if ABA_PLANILHA in wb.sheetnames:
        ws = wb[ABA_PLANILHA]
    else:
        ws = wb.active
        if not silencioso:
            print(f"  Aviso: aba '{ABA_PLANILHA}' nao encontrada. Usando aba ativa: {ws.title}")

    rows_out = []
    skipped_data = 0
    skipped_status = 0
    ignorados = defaultdict(int)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue

        status_raw = clean_str(row[COL["status"]]) if len(row) > COL["status"] else ""
        # status_raw = status_raw.replace(".", "")
        status = status_raw.upper()
        inv_rec = parse_date(row[COL["inv_receipt"]]) if len(row) > COL["inv_receipt"] else None

        if not inv_rec or inv_rec < DATA_CORTE:
            skipped_data += 1            
            continue
        else:
            if not silencioso and row[23] == "BALENCIAGA":
                if row[4] != "WAITING ARRIVAL":
                  print("Invoice ", contador, ": ", row[1])
                  contador = contador + 1

        if status not in status_filtro:
            skipped_status += 1
            ignorados[status or "(vazio)"] += 1
            continue

        eta = parse_date(row[COL["eta"]]) if len(row) > COL["eta"] else None
        ship = clean_str(row[COL["shipment"]])
        inv = clean_str(row[COL["invoice"]])
        boxes = int(row[COL["boxes"]]) if isinstance(row[COL["boxes"]], (int, float)) else 0
        items = int(row[COL["items"]]) if isinstance(row[COL["items"]], (int, float)) else 0
        loc = clean_str(row[COL["loc"]]) or "-"
        channel = clean_str(row[COL["channel"]])
        status_lead = clean_str(row[COL["status_lead"]])
        brand = clean_str(row[COL["brand"]]).upper()

        rows_out.append([
            brand, ship, boxes, items, status, loc, channel, status_lead,
            inv_rec.strftime("%Y-%m-%d"),
            eta.strftime("%Y-%m-%d") if eta else None,
            val_lt(row[COL["lt_ig"]]),
            val_lt(row[COL["lt_ai"]]),
            val_lt(row[COL["lt_ic"]]),
            val_lt(row[COL["lt_cn"]]),
            val_lt(row[COL["lt_np"]]),
            inv,
        ])

    if not silencioso:
        print(f"  Linhas validas: {len(rows_out)}")
        print(f"  Ignoradas por data (< {DATA_CORTE.strftime('%d/%m/%Y')}): {skipped_data}")
        print(f"  Ignoradas por status invalido: {skipped_status}")
        if ignorados:
            print("    Status ignorados encontrados:")
            for k, v in sorted(ignorados.items(), key=lambda x: -x[1])[:10]:
                print(f"      {k}: {v}")

    return rows_out


def imprimir_resumo(rows):
    by_brand = defaultdict(lambda: {"linhas": 0, "invoices": set(), "boxes": 0, "items": 0})
    for r in rows:
        b = r[0]
        by_brand[b]["linhas"] += 1
        if r[15]:
            by_brand[b]["invoices"].add(r[15])
        by_brand[b]["boxes"] += r[2]
        by_brand[b]["items"] += r[3]

    print("\n=== RESUMO ===")
    marcas = sorted(by_brand) if by_brand else []
    for brand in marcas:
        d = by_brand[brand]
        print(f"  {brand}: invoices={len(d['invoices'])} boxes={d['boxes']:,} items={d['items']:,}")
    print(f"  TOTAL: {len(rows):,} linhas")
    faltando = set(BRANDS_ESPERADAS) - set(marcas)
    if faltando:
        print(f"  Aviso: nenhuma linha encontrada para: {sorted(faltando)}")


MESES_ABREV = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr", "05": "May", "06": "Jun",
    "07": "Jul", "08": "Aug", "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}


def computar_datas(rows):
    """Extrai todas as datas (inv_receipt e eta) das linhas para determinar
    o intervalo real coberto pelos dados."""
    datas = []
    for r in rows:
        if r[8]:
            datas.append(datetime.datetime.strptime(r[8], "%Y-%m-%d"))
        if r[9]:
            datas.append(datetime.datetime.strptime(r[9], "%Y-%m-%d"))
    return datas


def computar_months(datas):
    """Retorna lista ordenada de meses (YYYY-MM) presentes nos dados e o
    dicionario de rotulos (Apr/26, May/26, ...) correspondente."""
    meses = sorted({d.strftime("%Y-%m") for d in datas})
    labels = {m: f"{MESES_ABREV[m[5:7]]}/{m[2:4]}" for m in meses}
    return meses, labels


def computar_weeks(datas):
    """Gera as semanas (segunda a domingo, padrao ISO) que cobrem o
    intervalo real de datas. Retorna (dict WEEK_RANGES, lista de opcoes
    [(chave, rotulo), ...]) prontos para injetar no HTML/JS."""
    if not datas:
        return {}, []

    dmin, dmax = min(datas), max(datas)
    inicio = dmin - datetime.timedelta(days=dmin.weekday())  # segunda-feira

    week_ranges = {}
    opcoes = []
    cursor = inicio
    while cursor <= dmax:
        fim = cursor + datetime.timedelta(days=6)
        iso_ano, iso_semana, _ = cursor.isocalendar()
        chave = f"{iso_ano}-W{iso_semana:02d}"
        week_ranges[chave] = [cursor.strftime("%Y-%m-%d"), fim.strftime("%Y-%m-%d")]

        if cursor.month == fim.month:
            rotulo = f"W{iso_semana} · {MESES_ABREV[f'{cursor.month:02d}']} {cursor.day}–{fim.day}"
        else:
            rotulo = (f"W{iso_semana} · {MESES_ABREV[f'{cursor.month:02d}']} {cursor.day} – "
                      f"{MESES_ABREV[f'{fim.month:02d}']} {fim.day}")
        opcoes.append((chave, rotulo))
        cursor += datetime.timedelta(days=7)

    return week_ranges, opcoes


# ══════════════════════════════════════════════════════════════════════════
# TEMPLATE HTML / CSS / JS DO DASHBOARD
# (gerado a partir do dashboard em producao — nao editar manualmente)
# ══════════════════════════════════════════════════════════════════════════

CSS_TEMPLATE = """@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;700;900&family=DM+Mono:wght@400;500&display=swap');
:root{
  --line:#1e2d45;--muted:#6b7f9e;--text:#f0f4fc;
  --blue:#3b82f6;--sky:#93c5fd;--orange:#f59e0b;--red:#ef4444;
  --violet:#818cf8;--green:#22c55e;--pink:#f472b6;
}
*{box-sizing:border-box;margin:0;padding:0}
html,body{background:radial-gradient(ellipse at 18% -8%,#1a3a6b 0%,#090f1c 40%,#070915 100%);font-family:'DM Sans',sans-serif;color:var(--text);min-height:100vh}
.header{padding:22px 36px 16px;display:flex;justify-content:space-between;align-items:flex-end;border-bottom:1px solid var(--line)}
.eyebrow{font-size:9px;color:#5b7db5;letter-spacing:5px;text-transform:uppercase;font-weight:700;margin-bottom:7px}
h1{font-size:21px;font-weight:900;letter-spacing:-.4px;color:#e8f0fc}
h1 span{color:var(--sky)}
.sub{color:var(--muted);font-size:11px;margin-top:4px}
.badge{display:inline-flex;align-items:center;gap:6px;background:rgba(59,130,246,.1);border:1px solid rgba(59,130,246,.25);border-radius:6px;padding:3px 10px;font-size:9px;font-weight:700;color:#93c5fd;letter-spacing:1px;text-transform:uppercase;margin-top:6px}
.badge span{width:6px;height:6px;border-radius:50%;background:#3b82f6;box-shadow:0 0 8px #3b82f6}
.hright{text-align:right;font-size:10px;color:var(--muted);font-family:'DM Mono',monospace;line-height:1.7}
.hright strong{color:#dbeafe;font-size:11px}
/* FILTER BAR */
.fbar{padding:10px 36px;background:rgba(10,16,30,.9);border-bottom:1px solid var(--line);display:flex;align-items:center;gap:14px;flex-wrap:wrap}
.flabel{font-size:8.5px;color:#4a6080;letter-spacing:2px;text-transform:uppercase;font-weight:700;white-space:nowrap}
.fmode{display:flex;background:#0a1428;border:1px solid var(--line);border-radius:7px;overflow:hidden}
.fmbtn{padding:5px 12px;font-size:9px;font-weight:700;cursor:pointer;color:#4a6080;letter-spacing:.5px;text-transform:uppercase;border:none;background:transparent;transition:all .15s;font-family:'DM Sans',sans-serif}
.fmbtn.active{background:#1a2e4a;color:#93c5fd}
.finputs{display:flex;align-items:center;gap:8px}
.finputs label{font-size:9px;color:#4a6080;font-weight:600}
input[type=date]{background:#0a1428;border:1px solid var(--line);color:#93c5fd;border-radius:6px;padding:5px 9px;font-size:10px;font-family:'DM Mono',monospace;outline:none;cursor:pointer}
input[type=date]::-webkit-calendar-picker-indicator{filter:invert(.4) sepia(1) saturate(2) hue-rotate(190deg);cursor:pointer}
.wsel{background:#0a1428;border:1px solid var(--line);color:#93c5fd;border-radius:6px;padding:5px 9px;font-size:10px;font-family:'DM Mono',monospace;outline:none;cursor:pointer}
.fclear{font-size:9px;color:#3a5270;cursor:pointer;padding:4px 8px;border-radius:5px;border:1px solid transparent;transition:all .15s;background:none;font-family:'DM Sans',sans-serif;font-weight:700}
.fclear:hover{color:#93c5fd;border-color:var(--line)}
.fcount{font-size:9.5px;color:#4a6080;font-family:'DM Mono',monospace}
.fcount strong{color:#93c5fd}
/* TABS */
.tabs{padding:0 36px;display:flex;gap:4px;border-bottom:1px solid var(--line)}
.tab{padding:10px 22px;border-radius:10px 10px 0 0;font-size:10.5px;font-weight:700;cursor:pointer;letter-spacing:.6px;text-transform:uppercase;border:1px solid transparent;border-bottom:none;position:relative;bottom:-1px;color:var(--muted);background:transparent;transition:all .15s}
.tab:hover{color:#c8daf5;background:rgba(255,255,255,.03)}
.tab.active{background:#0f1c30;border-color:var(--line);color:#e8f0fc}
.tab.active[data-b="balenciaga"]{color:#93c5fd;border-top:2px solid #93c5fd}
.tab.active[data-b="bottega"]{color:#60a5fa;border-top:2px solid #60a5fa}
.tab.active[data-b="ysl"]{color:#3b82f6;border-top:2px solid #3b82f6}
.tab.active[data-b="gucci"]{color:#93c5fd;border-top:2px solid #1d4ed8}
.tab.active[data-b="total"]{color:#22d3ee;border-top:2px solid #22d3ee}
.tdot{display:inline-block;width:5px;height:5px;border-radius:50%;margin-right:7px;vertical-align:middle}
.page{display:none;padding:20px 36px 48px;animation:fi .18s ease}
.page.active{display:block}
@keyframes fi{from{opacity:0;transform:translateY(3px)}to{opacity:1;transform:none}}
/* STATUS BAR */
.sbar{display:flex;align-items:center;gap:10px;background:rgba(15,24,44,.8);border:1px solid var(--line);border-radius:9px;padding:10px 16px;margin-bottom:16px;flex-wrap:wrap}
.spill{display:inline-flex;align-items:center;gap:6px;padding:4px 11px;border-radius:5px;font-size:9.5px;font-weight:700;letter-spacing:.5px}
.spill.transit{background:rgba(59,130,246,.1);color:#93c5fd;border:1px solid rgba(59,130,246,.2)}
.spill.delayed{background:rgba(239,68,68,.1);color:#f87171;border:1px solid rgba(239,68,68,.2)}
.spill.ok{background:rgba(34,197,94,.1);color:#6ee7b7;border:1px solid rgba(34,197,94,.2)}
.spill .sdot{width:5px;height:5px;border-radius:50%}
.sdiv{width:1px;height:20px;background:var(--line)}
.dbtn{display:inline-flex;align-items:center;gap:7px;background:rgba(239,68,68,.1);border:1px solid rgba(239,68,68,.3);color:#fca5a5;border-radius:7px;padding:5px 13px;font-size:9.5px;font-weight:700;cursor:pointer;text-transform:uppercase;transition:all .18s;margin-left:auto;font-family:'DM Sans',sans-serif}
.dbtn:hover{background:rgba(239,68,68,.18);color:#fecaca}
.dbtn .ddot{width:6px;height:6px;border-radius:50%;background:#ef4444;box-shadow:0 0 8px rgba(239,68,68,.6)}
/* SEC */
.sec{display:flex;align-items:center;gap:10px;margin:20px 0 11px;font-size:8.5px;color:#dbeafe;letter-spacing:4.5px;text-transform:uppercase;font-weight:800}
.sec:after{content:"";height:1px;background:var(--line);flex:1}
/* KPI */
.krow{display:grid;grid-template-columns:repeat(4,1fr);gap:11px;margin-bottom:4px}
.kpi{background:rgba(15,24,44,.95);border:1px solid var(--line);border-radius:12px;padding:15px 17px;position:relative;overflow:hidden}
.ka{width:2.5px;height:100%;position:absolute;left:0;top:0;border-radius:12px 0 0 12px}
.kl{font-size:8px;color:#e8f0fc;letter-spacing:2.5px;text-transform:uppercase;font-weight:800;margin-bottom:7px}
.kv{font-size:28px;font-weight:900;line-height:1;letter-spacing:-1px;font-family:'DM Mono',monospace}
.ks{font-size:10px;color:#3d5270;margin-top:5px}
/* PIPELINE */
.pwrap{background:rgba(12,19,36,.95);border:1px solid var(--line);border-radius:12px;overflow:hidden;margin-bottom:4px}
.phdr{display:grid;border-bottom:1px solid var(--line)}
.phdr-cell{padding:6px 13px;border-right:1px solid var(--line);font-size:7px;color:#dbeafe;text-transform:uppercase;letter-spacing:1.5px;font-weight:700}
.phdr-cell:last-child{border-right:none}
.plevel{display:grid;border-bottom:1px solid rgba(30,45,69,.5)}
.plevel:last-child{border-bottom:none}
.plbl{display:flex;align-items:center;padding:0 13px;border-right:1px solid rgba(30,45,69,.4);background:rgba(8,14,24,.4);min-height:52px;min-width:72px}
.plbl span{font-size:7.5px;color:#2a3f5f;text-transform:uppercase;letter-spacing:1.5px;font-weight:700;white-space:nowrap}
.pcell{padding:10px 13px;border-right:1px solid rgba(30,45,69,.4);position:relative;min-height:52px}
.pcell:last-child{border-right:none}
.pcval{font-size:20px;font-weight:900;font-family:'DM Mono',monospace;line-height:1}
.pcpct{font-size:8.5px;color:var(--muted);margin-top:3px}
.pcbar{height:2px;border-radius:2px;margin-top:7px}
.pcarr{position:absolute;right:-7px;top:50%;transform:translateY(-50%);font-size:10px;color:#1e2d45;z-index:2}
/* GRID */
.grid{display:grid;grid-template-columns:repeat(12,1fr);gap:13px}
.s12{grid-column:span 12}.s8{grid-column:span 8}.s6{grid-column:span 6}
.s5{grid-column:span 5}.s4{grid-column:span 4}.s3{grid-column:span 3}
.card{background:rgba(12,19,36,.92);border:1px solid var(--line);border-radius:12px;padding:15px 17px;box-shadow:0 8px 28px rgba(0,0,0,.25)}
.card h2{font-size:9px;margin:0 0 13px;text-transform:uppercase;letter-spacing:2px;color:#f0f4fc;display:flex;align-items:center;gap:7px;font-weight:800}
.cdot{width:5px;height:5px;border-radius:50%;flex-shrink:0}
.ch{position:relative;width:100%}
.h280{height:280px}.h200{height:200px}
svg{width:100%;height:100%;overflow:visible}
.gl{stroke:#131e30;stroke-width:1}
.ax{fill:#2e4060;font-size:9.5px;font-family:'DM Mono',monospace}
.legend{display:flex;gap:12px;align-items:center;justify-content:center;margin-top:8px;color:#4a6080;font-size:9.5px;font-weight:600;flex-wrap:wrap}
.li{display:flex;align-items:center;gap:4px}
.sw{display:inline-block;height:2.5px;width:14px;border-radius:3px}
.pills{display:flex;gap:5px;margin-bottom:8px;flex-wrap:wrap}
.pill{font-size:9px;padding:3px 9px;border-radius:5px;border:1px solid #1e3050;color:#4a6a9a;background:#0c1422;cursor:pointer;font-weight:700;transition:all .15s}
.pill:hover{border-color:var(--blue);color:#93c5fd}
.pill.active{background:#112244;color:#dbeafe;border-color:#2a5bb0}
.blist{display:flex;flex-direction:column;gap:6px}
.brow{display:flex;align-items:center;gap:8px}
.bnm{font-size:9.5px;color:#6b8ab0;width:170px;flex-shrink:0;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-weight:500}
.bbg{flex:1;height:15px;background:#0a1020;border-radius:3px;overflow:hidden}
.bfill{height:100%;border-radius:3px}
.bval{font-size:9px;color:var(--muted);width:30px;text-align:right;font-family:'DM Mono',monospace;flex-shrink:0}
.frow{display:flex;align-items:center;gap:8px;margin-bottom:7px}
.fflbl{font-size:9.5px;color:#5a7090;width:110px;text-align:right;font-weight:600;flex-shrink:0}
.ffbg{flex:1;height:21px;background:#0a1020;border-radius:4px;overflow:hidden}
.ffbar{height:100%;border-radius:4px;display:flex;align-items:center;justify-content:flex-end;padding-right:7px;font-size:9.5px;font-weight:700;color:rgba(255,255,255,.85);font-family:'DM Mono',monospace}
.ffavg{font-size:9.5px;color:var(--muted);width:30px;text-align:right;font-family:'DM Mono',monospace;flex-shrink:0}
/* SLA */
.sla-wrap{display:grid;grid-template-columns:repeat(4,1fr);gap:11px;margin-top:4px}
.nfpod-wrap{display:grid;grid-template-columns:repeat(3,1fr);gap:11px;margin-top:4px}
.sla-card{background:rgba(10,16,32,.92);border:1px solid var(--line);border-radius:11px;padding:13px 10px;text-align:center;cursor:default;transition:border-color .2s}
.sla-card:hover{border-color:rgba(59,130,246,.3)}
.sttl{font-size:8.5px;color:#f0f4fc;letter-spacing:1.5px;text-transform:uppercase;font-weight:700;margin-bottom:3px;line-height:1.4}
.stgt{font-size:7.5px;color:#263545;margin-bottom:7px}
.sdn{position:relative;width:78px;height:78px;margin:0 auto 9px}
.sdn svg{position:absolute;top:0;left:0;width:100%;height:100%}
.sctr{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center;pointer-events:none}
.spct{font-size:17px;font-weight:900;font-family:'DM Mono',monospace;line-height:1}
.ssub{font-size:7.5px;color:#4a6080;margin-top:1px}
.scts{display:flex;justify-content:center;gap:8px;font-size:8.5px;font-family:'DM Mono',monospace}
.sbar2{height:5px;background:#080f1c;border-radius:3px;margin-top:6px;overflow:hidden}
.sfill{height:100%;border-radius:3px}
/* MODAL */
.moverlay{position:fixed;inset:0;background:rgba(2,6,18,.82);z-index:1000;display:none;align-items:center;justify-content:center;backdrop-filter:blur(4px);padding:24px}
.moverlay.open{display:flex}
.modal{background:#0c1830;border:1px solid #263552;border-radius:16px;width:100%;max-width:860px;max-height:88vh;display:flex;flex-direction:column;box-shadow:0 24px 64px rgba(0,0,0,.7)}
.mhdr{padding:20px 24px 16px;border-bottom:1px solid #1e2d45;display:flex;align-items:center;justify-content:space-between;flex-shrink:0}
.mtitle{font-size:13px;font-weight:900;color:#e8f0fc}
.mtitle span{color:#93c5fd}
.msub{font-size:10px;color:#4a6080;margin-top:3px}
.mclose{width:32px;height:32px;border-radius:8px;background:#1a2a42;border:1px solid #263552;color:#6b8ab0;cursor:pointer;display:flex;align-items:center;justify-content:center;font-size:16px;transition:all .15s;flex-shrink:0}
.mclose:hover{background:#22334e;color:#dbeafe}
.mbody{padding:20px 24px;overflow-y:auto;flex:1}
.scard{background:#0a1428;border:1px solid #1a2a42;border-radius:12px;padding:16px 18px;margin-bottom:14px}
.scard:last-child{margin-bottom:0}
.shdr{display:flex;align-items:center;gap:12px;margin-bottom:14px;flex-wrap:wrap}
.sid{font-size:13px;font-weight:900;font-family:'DM Mono',monospace;color:#dbeafe;letter-spacing:.5px}
.smeta{display:flex;gap:8px;flex-wrap:wrap}
.stag{font-size:8.5px;padding:3px 9px;border-radius:5px;font-weight:700;letter-spacing:.5px}
.stag.loc{background:rgba(59,130,246,.12);color:#93c5fd;border:1px solid rgba(59,130,246,.2)}
.stag.inv{background:rgba(129,140,248,.12);color:#a5b4fc;border:1px solid rgba(129,140,248,.2)}
.stag.br{background:rgba(59,130,246,.12);color:#93c5fd;border:1px solid rgba(59,130,246,.2)}
.mschart{display:flex;flex-direction:column;gap:7px}
.msrow{display:flex;align-items:center;gap:10px}
.mslbl{font-size:9px;color:#6b8ab0;width:90px;text-align:right;font-weight:600;flex-shrink:0}
.mstrack{flex:1;height:20px;background:#080e1c;border-radius:4px;position:relative;overflow:hidden}
.msfill{height:100%;border-radius:4px;display:flex;align-items:center;padding-left:8px;font-size:9px;font-weight:700;font-family:'DM Mono',monospace;color:rgba(255,255,255,.9);white-space:nowrap}
.msline{position:absolute;top:0;height:100%;width:1.5px;background:rgba(255,255,255,.15);z-index:1}
.msval{font-size:9px;width:36px;text-align:right;font-family:'DM Mono',monospace;flex-shrink:0;font-weight:700}
.msslal{font-size:8px;color:#3a5070;width:32px;flex-shrink:0}
/* TOOLTIP */
.tip{position:fixed;display:none;background:#060d1a;border:1px solid #1e2d45;color:#d5e3f5;border-radius:8px;padding:9px 12px;font-size:10.5px;z-index:9999;pointer-events:none;box-shadow:0 10px 28px rgba(0,0,0,.6);max-width:220px;line-height:1.55;font-family:'DM Sans',sans-serif}
@media(max-width:1200px){.krow{grid-template-columns:repeat(2,1fr)}.s4,.s5,.s6,.s8{grid-column:span 12}.sla-wrap{grid-template-columns:repeat(2,1fr)}.nfpod-wrap{grid-template-columns:1fr 1fr}}
@media(max-width:700px){.krow{grid-template-columns:1fr 1fr}.header{flex-direction:column}.page{padding:14px 16px 32px}.tabs{padding:0 12px;overflow-x:auto}.tab{white-space:nowrap}.fbar{padding:10px 16px}.s3{grid-column:span 12}}"""

HTML_BODY_TEMPLATE = """
<div class="header">
  <div>
    <div class="eyebrow">Kering Group · Import Operations · Executive Dashboard</div>
    <h1>Process <span>Tracker</span> — All Open</h1>
    <div class="sub">Balenciaga · Bottega Veneta · YSL · Gucci · Delivered · Delivery Scheduled · Waiting Cargo · Waiting CC · Waiting Delivery · Waiting NF</div>
    <div class="badge"><span></span>Status: Delivered · Del. Scheduled · W. Cargo · W. CC · W. Del. Schedule · W. NF</div>
  </div>
  <div class="hright">Data range<br><strong>{DATA_RANGE_LABEL}</strong><br><span style="margin-top:3px;display:block">Last update &nbsp;<strong>{LAST_UPDATE_LABEL}</strong></span></div>
</div>

<div class="fbar">
  <div class="flabel">Filter by</div>
  <div class="fmode">
    <button class="fmbtn active" id="mode-eta" onclick="setFMode('eta')">ETA Date</button>
    <button class="fmbtn" id="mode-receipt" onclick="setFMode('receipt')">Invoice Receipt</button>
  </div>
  <div class="finputs">
    <label>From</label><input type="date" id="f-from" onchange="applyFilter()"/>
    <label>To</label><input type="date" id="f-to" onchange="applyFilter()"/>
  </div>
  <div style="display:flex;align-items:center;gap:8px">
    <label style="font-size:9px;color:#4a6080;font-weight:600">Week</label>
    <select class="wsel" id="f-week" onchange="applyWeek()">
      <option value="">— All weeks —</option>
      {WEEK_OPTIONS_HTML}
    </select>
  </div>
  <button class="fclear" onclick="clearFilter()">↺ Clear</button>
  <div class="fcount" id="fcount"></div>
</div>

<div class="tabs">
  <div class="tab active" data-b="balenciaga" onclick="setTab('balenciaga',this)"><span class="tdot" style="background:#93c5fd"></span>Balenciaga</div>
  <div class="tab" data-b="bottega" onclick="setTab('bottega',this)"><span class="tdot" style="background:#60a5fa"></span>Bottega Veneta</div>
  <div class="tab" data-b="ysl" onclick="setTab('ysl',this)"><span class="tdot" style="background:#3b82f6"></span>Saint Laurent</div>
  <div class="tab" data-b="gucci" onclick="setTab('gucci',this)"><span class="tdot" style="background:#1d4ed8"></span>Gucci</div>
  <div class="tab" data-b="total" onclick="setTab('total',this)"><span class="tdot" style="background:#22d3ee"></span>Total</div>
</div>


<div class="page active" id="page-balenciaga">
  <div id="sbar-balenciaga" class="sbar"></div>
  <div class="sec">Overview</div>
  <div class="krow" id="kpis-balenciaga"></div>
  <div class="sec">Pipeline Stage</div>
  <div class="pwrap" id="pipe-balenciaga"></div>
  <div class="grid" style="margin-top:13px">
    <div class="card s8"><h2><span class="cdot" style="background:#93c5fd"></span>Monthly Avg Lead Time — days per Milestone</h2>
      <div class="pills" id="pills-balenciaga"><span class="pill active" data-mode="all">All</span><span class="pill" data-mode="lt_inv_gl">Invoice→GL</span><span class="pill" data-mode="lt_arr_id">ARR→ID Reg</span><span class="pill" data-mode="lt_cc_nf">CC→NF</span></div>
      <div class="ch h280" id="c-balenciaga-trend"></div><div class="legend" id="lg-balenciaga"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#f59e0b"></span>Avg Lead Time by Milestone</h2><div id="f-balenciaga"></div></div>
  </div>
  <div class="grid" style="margin-top:13px">
    <div class="card s3"><h2><span class="cdot" style="background:#818cf8"></span>Stage</h2><div class="blist" id="stage-balenciaga"></div></div>
    <div class="card s5"><h2><span class="cdot" style="background:#f59e0b"></span>Status</h2><div class="blist" id="status-balenciaga"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#3b82f6"></span>PENDING INVOICES VOLUME BY MONTH</h2><div class="ch h200" id="vol-balenciaga"></div></div>
  </div>
  <!-- <div class="sec">SLA Compliance</div><div class="sla-wrap" id="sla-balenciaga"></div> -->
  <!-- <div class="sec">NF → POD by State &nbsp;<span style="color:#1e3050;font-size:8px;letter-spacing:1px">(SAO 2d · RIO 3d · Others 5d)</span></div> -->
  <!--<div class="nfpod-wrap" id="nfpod-balenciaga"></div> -->
</div>
<div class="page" id="page-bottega">
  <div id="sbar-bottega" class="sbar"></div>
  <div class="sec">Overview</div>
  <div class="krow" id="kpis-bottega"></div>
  <div class="sec">Pipeline Stage</div>
  <div class="pwrap" id="pipe-bottega"></div>
  <div class="grid" style="margin-top:13px">
    <div class="card s8"><h2><span class="cdot" style="background:#60a5fa"></span>Monthly Avg Lead Time — days per Milestone</h2>
      <div class="pills" id="pills-bottega"><span class="pill active" data-mode="all">All</span><span class="pill" data-mode="lt_inv_gl">Invoice→GL</span><span class="pill" data-mode="lt_arr_id">ARR→ID Reg</span><span class="pill" data-mode="lt_cc_nf">CC→NF</span><span class="pill" data-mode="lt_nf_pod">NF→POD</span></div>
      <div class="ch h280" id="c-bottega-trend"></div><div class="legend" id="lg-bottega"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#f59e0b"></span>Avg Lead Time by Milestone</h2><div id="f-bottega"></div></div>
  </div>
  <div class="grid" style="margin-top:13px">
    <div class="card s3"><h2><span class="cdot" style="background:#818cf8"></span>Stage</h2><div class="blist" id="stage-bottega"></div></div>
    <div class="card s5"><h2><span class="cdot" style="background:#f59e0b"></span>Status</h2><div class="blist" id="status-bottega"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#3b82f6"></span>PENDING INVOICES VOLUME BY MONTH</h2><div class="ch h200" id="vol-bottega"></div></div>
  </div>
  <!-- <div class="sec">SLA Compliance</div><div class="sla-wrap" id="sla-bottega"></div> -->
  <!-- <div class="sec">NF → POD by State &nbsp;<span style="color:#1e3050;font-size:8px;letter-spacing:1px">(SAO 2d · RIO 3d · Others 5d)</span></div> -->
  <!-- <div class="nfpod-wrap" id="nfpod-bottega"></div> -->
</div>
<div class="page" id="page-ysl">
  <div id="sbar-ysl" class="sbar"></div>
  <div class="sec">Overview</div>
  <div class="krow" id="kpis-ysl"></div>
  <div class="sec">Pipeline Stage</div>
  <div class="pwrap" id="pipe-ysl"></div>
  <div class="grid" style="margin-top:13px">
    <div class="card s8"><h2><span class="cdot" style="background:#3b82f6"></span>Monthly Avg Lead Time — days per Milestone</h2>
      <div class="pills" id="pills-ysl"><span class="pill active" data-mode="all">All</span><span class="pill" data-mode="lt_inv_gl">Invoice→GL</span><span class="pill" data-mode="lt_arr_id">ARR→ID Reg</span><span class="pill" data-mode="lt_cc_nf">CC→NF</span><span class="pill" data-mode="lt_nf_pod">NF→POD</span></div>
      <div class="ch h280" id="c-ysl-trend"></div><div class="legend" id="lg-ysl"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#f59e0b"></span>Avg Lead Time by Milestone</h2><div id="f-ysl"></div></div>
  </div>
  <div class="grid" style="margin-top:13px">
    <div class="card s3"><h2><span class="cdot" style="background:#818cf8"></span>Stage</h2><div class="blist" id="stage-ysl"></div></div>
    <div class="card s5"><h2><span class="cdot" style="background:#f59e0b"></span>Status</h2><div class="blist" id="status-ysl"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#3b82f6"></span>PENDING INVOICES VOLUME BY MONTH</h2><div class="ch h200" id="vol-ysl"></div></div>
  </div>
  <!-- <div class="sec">SLA Compliance</div><div class="sla-wrap" id="sla-ysl"></div> -->
  <!-- <div class="sec">NF → POD by State &nbsp;<span style="color:#1e3050;font-size:8px;letter-spacing:1px">(SAO 2d · RIO 3d · Others 5d)</span></div> -->
  <!-- <div class="nfpod-wrap" id="nfpod-ysl"></div> -->
</div>
<div class="page" id="page-gucci">
  <div id="sbar-gucci" class="sbar"></div>
  <div class="sec">Overview</div>
  <div class="krow" id="kpis-gucci"></div>
  <div class="sec">Pipeline Stage</div>
  <div class="pwrap" id="pipe-gucci"></div>
  <div class="grid" style="margin-top:13px">
    <div class="card s8"><h2><span class="cdot" style="background:#1d4ed8"></span>Monthly Avg Lead Time — days per Milestone</h2>
      <div class="pills" id="pills-gucci"><span class="pill active" data-mode="all">All</span><span class="pill" data-mode="lt_inv_gl">Invoice→GL</span><span class="pill" data-mode="lt_arr_id">ARR→ID Reg</span><span class="pill" data-mode="lt_cc_nf">CC→NF</span><span class="pill" data-mode="lt_nf_pod">NF→POD</span></div>
      <div class="ch h280" id="c-gucci-trend"></div><div class="legend" id="lg-gucci"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#f59e0b"></span>Avg Lead Time by Milestone</h2><div id="f-gucci"></div></div>
  </div>
  <div class="grid" style="margin-top:13px">
    <div class="card s3"><h2><span class="cdot" style="background:#818cf8"></span>Stage</h2><div class="blist" id="stage-gucci"></div></div>
    <div class="card s5"><h2><span class="cdot" style="background:#f59e0b"></span>Status</h2><div class="blist" id="status-gucci"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#3b82f6"></span>PENDING INVOICES VOLUME BY MONTH</h2><div class="ch h200" id="vol-gucci"></div></div>
  </div>
  <!-- <div class="sec">SLA Compliance</div><div class="sla-wrap" id="sla-gucci"></div> -->
  <!-- <div class="sec">NF → POD by State &nbsp;<span style="color:#1e3050;font-size:8px;letter-spacing:1px">(SAO 2d · RIO 3d · Others 5d)</span></div> -->
  <!-- <div class="nfpod-wrap" id="nfpod-gucci"></div> -->
</div>
<div class="page" id="page-total">
  <div id="sbar-total" class="sbar"></div>
  <div class="sec">Overview</div>
  <div class="krow" id="kpis-total"></div>
  <div class="sec">Pipeline Stage</div>
  <div class="pwrap" id="pipe-total"></div>
  <div class="grid" style="margin-top:13px">
    <div class="card s8"><h2><span class="cdot" style="background:#22d3ee"></span>Monthly Avg Lead Time — days per Milestone</h2>
      <div class="pills" id="pills-total"><span class="pill active" data-mode="all">All</span><span class="pill" data-mode="lt_inv_gl">Invoice→GL</span><span class="pill" data-mode="lt_arr_id">ARR→ID Reg</span><span class="pill" data-mode="lt_cc_nf">CC→NF</span><span class="pill" data-mode="lt_nf_pod">NF→POD</span></div>
      <div class="ch h280" id="c-total-trend"></div><div class="legend" id="lg-total"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#f59e0b"></span>Avg Lead Time by Milestone</h2><div id="f-total"></div></div>
  </div>
  <div class="grid" style="margin-top:13px">
    <div class="card s3"><h2><span class="cdot" style="background:#818cf8"></span>Stage</h2><div class="blist" id="stage-total"></div></div>
    <div class="card s5"><h2><span class="cdot" style="background:#f59e0b"></span>Status</h2><div class="blist" id="status-total"></div></div>
    <div class="card s4"><h2><span class="cdot" style="background:#3b82f6"></span>PENDING INVOICES VOLUME BY MONTH</h2><div class="ch h200" id="vol-total"></div></div>
  </div>
  <!-- <div class="sec">SLA Compliance</div><div class="sla-wrap" id="sla-total"></div> -->
  <!-- <div class="sec">NF → POD by State &nbsp;<span style="color:#1e3050;font-size:8px;letter-spacing:1px">(SAO 2d · RIO 3d · Others 5d)</span></div> -->
  <!-- <div class="nfpod-wrap" id="nfpod-total"></div> -->
</div>

<div class="moverlay" id="moverlay" onclick="if(event.target===this)closeMod()">
  <div class="modal">
    <div class="mhdr">
      <div><div class="mtitle" id="mtitle"></div><div class="msub" id="msub"></div></div>
      <button class="mclose" onclick="closeMod()">✕</button>
    </div>
    <div class="mbody" id="mbody"></div>
  </div>
</div>
<div class="tip" id="tip"></div>

"""

JS_AFTER_DATA_TEMPLATE = """

// ── CONFIGURAÇÃO ─────────────────────────────────────────────
const WEEK_RANGES={WEEK_RANGES_JSON};
const STAGE_C={'Waiting Cargo Attendance':'#fb923c','Waiting ID Reg':'#fbbf24','Waiting CC':'#fde047','Waiting NF':'#f472b6','Waiting POD':'#34d399', 'Delivery Scheduled':'#21c3b0'};
const STAGE_ORDER=['Waiting Cargo Attendance','Waiting ID Reg','Waiting CC','Waiting NF','Waiting POD', 'Delivery Scheduled'];
const STATUS_C={'WAITING ID REGISTER':'#fbbf24','WAITING NF':'#f472b6','WAITING DELIVERY SCHEDULE':'#34d399','DELIVERY SCHEDULED':'#21c3b0','WAITING ARRIVAL':'#60a5fa','WAITING CARGO ATTENDANCE':'#fb923c','WAITING IBAMA':'#f87171','WAITING CUSTOMS CLEARANCE':'#fde047'};
const NF_ST=new Set(['DELIVERED','DELIVERY SCHEDULED','WAITING CARGO ATTENDANCE','WAITING CUSTOMS CLEARANCE','WAITING DELIVERY SCHEDULE','WAITING NF']);
const BRAND_C={'BALENCIAGA':'#93c5fd','BOTTEGA':'#60a5fa','GUCCI':'#1d4ed8','YSL':'#3b82f6'};
const BRAND_KEY={'BALENCIAGA':'balenciaga','BOTTEGA':'bottega','GUCCI':'gucci','YSL':'ysl'};
const META={lt_inv_gl:{label:'Invoice→GL',color:'#38bdf8',sla:7},lt_arr_id:{label:'ARR→ID Reg',color:'#f472b6',sla:2},lt_id_cc_ry:{label:'ID→CC (R/Y)',color:'#fbbf24',sla:1},lt_cc_nf:{label:'CC→NF',color:'#34d399',sla:2},lt_nf_pod:{label:'NF→POD',color:'#fb923c',sla:3}};
const MONTHS={MONTHS_JSON};
const MLABELS={MLABELS_JSON};

// ── ESTADO ───────────────────────────────────────────────────
let fMode='eta';
let fFrom=null, fTo=null;
let DATA={};
let DATA_ALL_STS={};
let tModes={balenciaga:'all',bottega:'all',ysl:'all',gucci:'all',total:'all'};

// ── HELPERS ──────────────────────────────────────────────────
function isN(x){return x!=null&&typeof x==='number'&&x>=0&&x<=120;}
function avN(arr){return arr.length?Math.round(arr.reduce((a,b)=>a+b,0)/arr.length*10)/10:null;}
function getStage(s){
  if(s==='WAITING CARGO ATTENDANCE') return 'Waiting Cargo Attendance';
  if(s==='WAITING ID REGISTER') return 'Waiting ID Reg';
  if(s==='WAITING CUSTOMS CLEARANCE') return 'Waiting CC';
  if(s==='WAITING NF') return 'Waiting NF';
  if(s==='WAITING DELIVERY SCHEDULE') return 'Waiting POD';
  if(s==='DELIVERY SCHEDULED') return 'Delivery Scheduled';
  //if(s==='WAITING GL') return 'Waiting GL';
  //if(s==='WAITING PRE ALERT') return 'Waiting Pre Alert';
  //if(s==='WAITING ARRIVAL') return 'Waiting Arr';
}
function pc(p){
  if(p>=90)return'#22c55e';if(p>=80)return'#93c5fd';if(p>=70)return'#60a5fa';
  if(p>=60)return'#3b82f6';if(p>=50)return'#2563eb';if(p>=40)return'#1d4ed8';
  return'#1e3a8a';
}

// ── TOOLTIP ──────────────────────────────────────────────────
const tipEl=document.getElementById('tip');
function showTip(e,h){tipEl.innerHTML=h;tipEl.style.display='block';mvTip(e);}
function mvTip(e){tipEl.style.left=(e.clientX+13)+'px';tipEl.style.top=(e.clientY-8)+'px';}
function hideTip(){tipEl.style.display='none';}
function svgTips(el){el.querySelectorAll('[data-tip]').forEach(n=>{n.addEventListener('mousemove',e=>showTip(e,n.getAttribute('data-tip')));n.addEventListener('mouseleave',hideTip);});}

// ── COMPUTE ──────────────────────────────────────────────────
function compute(rs){
  if(!rs||!rs.length) return null;
  const brand=rs[0][0];
  const delayed=rs.filter(r=>r[7]==='DELAYED');
  const uniq=a=>new Set(a.filter(Boolean)).size;
  const ships=uniq(rs.filter(r=>r[1]!=='TO CONFIRM').map(r=>r[1]));
  const uinv=uniq(rs.map(r=>r[15]))||rs.length;
  const delayedInv=uniq(delayed.map(r=>r[15]))||delayed.length;
  const boxes=rs.reduce((s,r)=>s+r[2],0);
  const items=rs.reduce((s,r)=>s+r[3],0);
  // pipeline
  const pipe=[];
  for(const st of STAGE_ORDER){
    const sr=rs.filter(r=>getStage(r[4])===st);
    if(!sr.length) continue;
    pipe.push({l:st,cs:[uniq(sr.filter(r=>r[1]!=='TO CONFIRM').map(r=>r[1])),uniq(sr.map(r=>r[15]))||sr.length,sr.reduce((s,r)=>s+r[2],0),sr.reduce((s,r)=>s+r[3],0)],c:STAGE_C[st]});
  }
  // stages
  const stgM={};rs.forEach(r=>{const s=getStage(r[4]);stgM[s]=(stgM[s]||0)+1;});
  const stages=Object.entries(stgM).sort((a,b)=>b[1]-a[1]).map(([n,v])=>({n,v,c:STAGE_C[n]||'#3b82f6'}));
  // statuses
  const scM={};rs.forEach(r=>{scM[r[4]]=(scM[r[4]]||0)+1;});
  const statuses=Object.entries(scM).sort((a,b)=>b[1]-a[1]).slice(0,7).map(([n,v])=>({n,v,c:STATUS_C[n]||'#3b82f6'}));
  // avgs
  const af=(fi,ry=false)=>avN(rs.filter(r=>isN(r[fi])&&(!ry||['RED','YELLOW'].includes(r[6]))).map(r=>r[fi]));
  const avgs={lt_inv_gl:af(10),lt_arr_id:af(11),lt_id_cc_ry:af(12,true),lt_cc_nf:af(13),lt_nf_pod:af(14)};
  // vol
  const vol={};rs.forEach(r=>{if(r[8]){const ym=r[8].slice(0,7);vol[ym]=(vol[ym]||0)+1;}});
  // trend
  const tmap={};
  rs.forEach(r=>{
    if(!r[8])return;const ym=r[8].slice(0,7);
    if(!tmap[ym])tmap[ym]={};
    [[10,'lt_inv_gl'],[11,'lt_arr_id'],[13,'lt_cc_nf'],[14,'lt_nf_pod']].forEach(([fi,k])=>{
      if(isN(r[fi])){if(!tmap[ym][k])tmap[ym][k]=[];tmap[ym][k].push(r[fi]);}
    });
  });
  const trend={};
  Object.keys(tmap).sort().forEach(ym=>{trend[ym]={};Object.entries(tmap[ym]).forEach(([k,v])=>{trend[ym][k]=avN(v);});});
  // SLA
  const sc2=(fi,sla,ry=false)=>{
    const el=ry?rs.filter(r=>['RED','YELLOW'].includes(r[6])):rs;
    return{ok:el.filter(r=>isN(r[fi])&&r[fi]<=sla).length,br:el.filter(r=>isN(r[fi])&&r[fi]>sla).length,na:el.filter(r=>!isN(r[fi])).length};
  };
  const sla={lt_inv_gl:sc2(10,7),lt_arr_id:sc2(11,2),lt_id_cc_ry:sc2(12,1,true),lt_cc_nf:sc2(13,2)};
  // NF→POD
  const nfpod={SAO:{ok:0,br:0,sla:2},RIO:{ok:0,br:0,sla:3},OTHER:{ok:0,br:0,sla:5}};
  rs.forEach(r=>{
    if(!NF_ST.has(r[4])||!isN(r[14]))return;
    const lk=['SAO','RIO'].includes(r[5])?r[5]:'OTHER';
    r[14]<=nfpod[lk].sla?nfpod[lk].ok++:nfpod[lk].br++;
  });
  // delayed detail
  const dsh={};
  delayed.forEach(r=>{
    const s=r[1]||'N/A';
    if(!dsh[s])dsh[s]={invSet:new Set(),loc:new Set(),lts:[]};
    if(r[15])dsh[s].invSet.add(r[15]);dsh[s].loc.add(r[5]);dsh[s].lts.push(r);
  });
  const del_list=Object.entries(dsh).sort().map(([ship,d2])=>{
    const locs=[...d2.loc].sort().join(',');
    const nps=d2.loc.has('SAO')?2:d2.loc.has('RIO')?3:5;
    const ma=fi=>avN(d2.lts.map(r=>r[fi]).filter(v=>isN(v)));
    return{shipment:ship,invoices:d2.invSet.size||d2.lts.length,location:locs,
      milestones:[{key:'Invoice→GL',val:ma(10),sla:7,color:'#93c5fd'},{key:'ARR→ID Reg',val:ma(11),sla:2,color:'#60a5fa'},{key:'ID Reg→CC',val:ma(12),sla:1,color:'#3b82f6'},{key:'CC→NF',val:ma(13),sla:nps,color:'#2563eb'},{key:'NF→POD',val:ma(14),sla:nps,color:'#1d4ed8'}]};
  });
  const aanjo = rs.map(r => r[1]);
  return{color:BRAND_C[brand]||'#3b82f6',transit:uinv-delayedInv,delayed:delayedInv,ships,inv:uinv,boxes,items,pipeline:pipe,stages,statuses,avgs,vol,trend,sla,nfpod,del_list, aanjo};
}

// ── FILTER ───────────────────────────────────────────────────
function getRows(brand,src){
  src=src||ALL_ROWS;
  let rs=src.filter(r=>r[0]===brand);
  if(fFrom||fTo){
    const fi=fMode==='eta'?9:8;
    rs=rs.filter(r=>{
      const d=r[fi];if(!d)return false;
      if(fFrom&&d<fFrom)return false;
      if(fTo&&d>fTo)return false;
      return true;
    });
  }
  return rs;
}

// Função idêntica à getRows(), porém trabalha com a costante ALL_STS_ROWS (que inclui os status 'WAITING ARRIVAL', 'WAITING GL' e 'WAITING PRE ALERT')
function getRowsSTS(brand,src){
  src=src||ALL_STS_ROWS;
  let rs=src.filter(r=>r[0]===brand);
  if(fFrom||fTo){
    const fi=fMode==='eta'?9:8;
    rs=rs.filter(r=>{
      const d=r[fi];if(!d)return false;
      if(fFrom&&d<fFrom)return false;
      if(fTo&&d>fTo)return false;
      return true;
    });
  }
  return rs;
}

// Agrega os 4 numeros de "ja entregue" (shipments/invoices/boxes/items)
// a partir de um conjunto de linhas DELIVERED ja filtradas.
function delivStats(rs){
  const uniq=a=>new Set(a.filter(Boolean)).size;
  const ships=uniq(rs.filter(r=>r[1]!=='TO CONFIRM').map(r=>r[1]));
  const inv=uniq(rs.map(r=>r[15]))||rs.length;
  const boxes=rs.reduce((s,r)=>s+r[2],0);
  const items=rs.reduce((s,r)=>s+r[3],0);
  return{ships,inv,boxes,items};
}
function recompute(){
  ['BALENCIAGA','BOTTEGA','GUCCI','YSL'].forEach(b=>{
    const key=BRAND_KEY[b];
    DATA_ALL_STS[key]=compute(getRows(b, ALL_STS_ROWS));
    DATA[key]=compute(getRows(b))||{color:BRAND_C[b],transit:0,delayed:0,ships:0,inv:0,boxes:0,items:0,pipeline:[],stages:[],statuses:[],avgs:{},vol:{},trend:{},sla:{lt_inv_gl:{ok:0,br:0,na:0},lt_arr_id:{ok:0,br:0,na:0},lt_id_cc_ry:{ok:0,br:0,na:0},lt_cc_nf:{ok:0,br:0,na:0}},nfpod:{SAO:{ok:0,br:0,sla:2},RIO:{ok:0,br:0,sla:3},OTHER:{ok:0,br:0,sla:5}},del_list:[]};
    DATA[key].delivered=delivStats(getRows(b,DELIVERED_ROWS));
    console.log("DATA KEY ACIMA");
  });
  const totRows=['BALENCIAGA','BOTTEGA','GUCCI','YSL'].flatMap(b=>getRows(b));
  DATA.total=compute(totRows)||{color:'#22d3ee',transit:0,delayed:0,ships:0,inv:0,boxes:0,items:0,pipeline:[],stages:[],statuses:[],avgs:{},vol:{},trend:{},sla:{lt_inv_gl:{ok:0,br:0,na:0},lt_arr_id:{ok:0,br:0,na:0},lt_id_cc_ry:{ok:0,br:0,na:0},lt_cc_nf:{ok:0,br:0,na:0}},nfpod:{SAO:{ok:0,br:0,sla:2},RIO:{ok:0,br:0,sla:3},OTHER:{ok:0,br:0,sla:5}},del_list:[]};
  const totStsRows=['BALENCIAGA','BOTTEGA','GUCCI','YSL'].flatMap(b=>getRows(b, ALL_STS_ROWS));
  DATA_ALL_STS.total=compute(totStsRows);
  if(DATA.total) DATA.total.color='#22d3ee';
  const totDeliveredRows=['BALENCIAGA','BOTTEGA','GUCCI','YSL'].flatMap(b=>getRows(b,DELIVERED_ROWS));
  DATA.total.delivered=delivStats(totDeliveredRows);
}
function setFMode(m){
  fMode=m;
  document.getElementById('mode-eta').classList.toggle('active',m==='eta');
  document.getElementById('mode-receipt').classList.toggle('active',m==='receipt');
  document.getElementById('f-week').value='';
  applyFilter();
}
function applyWeek(){
  const w=document.getElementById('f-week').value;
  if(!w){clearFilter();return;}
  const [fr,to]=WEEK_RANGES[w]||[];
  if(fr){document.getElementById('f-from').value=fr;document.getElementById('f-to').value=to;}
  applyFilter();
}
function applyFilter(){
  fFrom=document.getElementById('f-from').value||null;
  fTo=document.getElementById('f-to').value||null;
  if(!fFrom&&!fTo){clearFilter();return;}
  recompute();
  const total=Object.values(DATA).reduce((s,d)=>s+(d.inv||0),0);
  const lbl=fMode==='eta'?'ETA':'Invoice Receipt';
  document.getElementById('fcount').innerHTML=`<strong>${total.toLocaleString()}</strong> processes · ${lbl}: ${fFrom||'—'} → ${fTo||'—'}`;
  renderAll();
}
function clearFilter(){
  fFrom=null;fTo=null;
  document.getElementById('f-from').value='';
  document.getElementById('f-to').value='';
  document.getElementById('f-week').value='';
  document.getElementById('fcount').innerHTML='';
  recompute();
  renderAll();
}

// ── TABS ─────────────────────────────────────────────────────
function setTab(name,el){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('page-'+name).classList.add('active');
  el.classList.add('active');
  renderAll();
}

// ── STATUS BAR ───────────────────────────────────────────────
function drawSbar(brand){
  const d=DATA[brand];const el=document.getElementById('sbar-'+brand);if(!el||!d)return;
  const hasD=d.del_list&&d.del_list.length>0;
  const totD=hasD?d.del_list.reduce((s,x)=>s+x.invoices,0):0;
  let h=`<div class="spill transit"><span class="sdot" style="background:#3b82f6"></span>INVOICES ON GOING &nbsp;<strong>${d.transit.toLocaleString()}</strong></div>`;
  if(hasD){
    h+=`<div class="sdiv"></div><div class="spill delayed"><span class="sdot" style="background:#ef4444"></span>INVOICES DELAYED &nbsp;<strong>${totD}</strong></div>
    <div class="sdiv"></div><span style="font-size:9.5px;color:#3d5270">Total: <strong style="color:#dbeafe">${d.inv.toLocaleString()}</strong></span>
    <button class="dbtn" onclick="openMod('${brand}')"><span class="ddot"></span>View Delayed Shipments</button>`;
  } else {
    h+=`<div class="sdiv"></div><div class="spill ok"><span class="sdot" style="background:#22c55e"></span>NO DELAYED INVOICES</div>
    <div class="sdiv"></div><span style="font-size:9.5px;color:#3d5270">Total: <strong style="color:#dbeafe">${d.inv.toLocaleString()}</strong></span>`;
  }
  el.innerHTML=h;
}

// ── KPIs ─────────────────────────────────────────────────────
function drawKpis(brand){
  const d=DATA[brand];const el=document.getElementById('kpis-'+brand);if(!el||!d)return;
  const dl=d.delivered||{ships:0,inv:0,boxes:0,items:0};
  el.innerHTML=`
    <div class="kpi"><div class="ka" style="background:#1d4ed8"></div><div class="kl">Open Shipments</div><div class="kv" style="color:#93c5fd">${d.ships.toLocaleString()}</div><div class="ks">Unique embarques</div></div>
    <div class="kpi"><div class="ka" style="background:#3b82f6"></div><div class="kl">Open Invoices</div><div class="kv" style="color:${d.color}">${d.inv.toLocaleString()}</div><div class="ks">N.º invoices abertas</div></div>
    <div class="kpi"><div class="ka" style="background:#60a5fa"></div><div class="kl">Open Boxes</div><div class="kv" style="color:#93c5fd">${d.boxes.toLocaleString()}</div><div class="ks">Caixas em processo</div></div>
    <div class="kpi"><div class="ka" style="background:#818cf8"></div><div class="kl">Open Items</div><div class="kv" style="color:#818cf8">${d.items.toLocaleString()}</div><div class="ks">Itens em processo</div></div>
    <div class="kpi"><div class="ka" style="background:#22c55e"></div><div class="kl">Delivered Shipments</div><div class="kv" style="color:#6ee7b7">${dl.ships.toLocaleString()}</div><div class="ks">Embarques entregues</div></div>
    <div class="kpi"><div class="ka" style="background:#22c55e"></div><div class="kl">Delivered Invoices</div><div class="kv" style="color:#6ee7b7">${dl.inv.toLocaleString()}</div><div class="ks">N.º invoices entregues</div></div>
    <div class="kpi"><div class="ka" style="background:#22c55e"></div><div class="kl">Delivered Boxes</div><div class="kv" style="color:#6ee7b7">${dl.boxes.toLocaleString()}</div><div class="ks">Caixas entregues</div></div>
    <div class="kpi"><div class="ka" style="background:#22c55e"></div><div class="kl">Delivered Items</div><div class="kv" style="color:#6ee7b7">${dl.items.toLocaleString()}</div><div class="ks">Itens entregues</div></div>`;
}

// ── PIPELINE ─────────────────────────────────────────────────
function drawPipe(brand){
  const d=DATA[brand];const el=document.getElementById('pipe-'+brand);if(!el||!d)return;
  const stages=d.pipeline;if(!stages||!stages.length){el.innerHTML='<div style="padding:16px;color:#dbeafe;font-size:10px">No data for current filter</div>';return;}
  const cols=stages.length;
  const totals=[0,1,2,3].map(i=>stages.reduce((s,st)=>s+st.cs[i],0));
  const lvlLbls=['Shipments','Invoices','Boxes','Items'];
  const lvlCols=['#93c5fd','#60a5fa','#a78bfa','#818cf8'];
  let h=`<div class="phdr" style="grid-template-columns:72px repeat(${cols},1fr)"><div class="phdr-cell" style="background:rgba(8,14,24,.6)">Level</div>`;
  stages.forEach(s=>{h+=`<div class="phdr-cell" style="background:rgba(8,14,24,.4)">${s.l}</div>`;});
  h+='</div>';
  [0,1,2,3].forEach(lv=>{
    h+=`<div class="plevel" style="grid-template-columns:72px repeat(${cols},1fr)"><div class="plbl"><span style="color:${lvlCols[lv]}">${lvlLbls[lv]}</span></div>`;
    stages.forEach((s,si)=>{
      const v=s.cs[lv];const tot=totals[lv]||1;const pct=tot>0?(v/tot*100).toFixed(1):'0';
      const isLast=si===stages.length-1;

      h+=`<div class="pcell"><div class="pcval" style="color:${s.c};font-size:${lv===0?18:lv===1?20:lv===2?16:15}px">${v.toLocaleString()}</div><div class="pcpct">${pct}%</div><div class="pcbar" style="background:${s.c};opacity:${lv===0?0.9:lv===1?0.75:lv===2?0.6:0.45}"></div>${!isLast?'<div class="pcarr">›</div>':''}</div>`;
    });
    h+='</div>';
  });
  el.innerHTML=h;
}

// ── TREND ────────────────────────────────────────────────────
function drawTrend(brand){
  const d=DATA[brand];const el=document.getElementById('c-'+brand+'-trend');if(!el||!d)return;
  const mode=tModes[brand];
  const w=el.clientWidth||700,h=el.clientHeight||260;
  const m={t:28,r:20,b:40,l:46};
  const base=['lt_inv_gl','lt_arr_id','lt_cc_nf'];
  const all=d.avgs&&d.avgs.lt_nf_pod!=null?[...base,'lt_nf_pod']:base;
  const keys=mode==='all'?all:[mode];
  let maxV=0;
  MONTHS.forEach(mo=>keys.forEach(k=>{const v=d.trend&&d.trend[mo]&&d.trend[mo][k];if(v!=null&&v>maxV)maxV=v;}));
  maxV=Math.ceil(maxV*1.22)||15;
  const gap=(w-m.l-m.r)/MONTHS.length;
  let svg=`<svg viewBox="0 0 ${w} ${h}">`;
  for(let i=0;i<=4;i++){const y=h-m.b-(h-m.t-m.b)*i/4;svg+=`<line class="gl" x1="${m.l}" y1="${y}" x2="${w-m.r}" y2="${y}"/><text class="ax" x="${m.l-7}" y="${y+4}" text-anchor="end">${Math.round(maxV*i/4)}d</text>`;}
  MONTHS.forEach((mo,i)=>{svg+=`<text class="ax" x="${m.l+i*gap+gap/2}" y="${h-4}" text-anchor="middle">${MLABELS[mo]}</text>`;});
  const allLabels=[];
  keys.forEach(k=>{
    const col=META[k]?META[k].color:'#93c5fd';
    const pts=MONTHS.map((mo,i)=>{const v=d.trend&&d.trend[mo]&&d.trend[mo][k];if(v==null)return null;return{x:m.l+i*gap+gap/2,y:h-m.b-(h-m.t-m.b)*v/maxV,v,mo};}).filter(Boolean);
    if(!pts.length)return;
    if(pts.length>=2){
      let path=`M ${pts[0].x} ${pts[0].y}`;
      for(let i=1;i<pts.length;i++){const cx=(pts[i-1].x+pts[i].x)/2;path+=` C ${cx} ${pts[i-1].y} ${cx} ${pts[i].y} ${pts[i].x} ${pts[i].y}`;}
      svg+=`<path d="${path}" fill="none" stroke="${col}" stroke-width="2.6"/><path d="${path} L ${pts[pts.length-1].x} ${h-m.b} L ${pts[0].x} ${h-m.b} Z" fill="${col}" fill-opacity="0.06"/>`;
    } else {
      svg+=`<line x1="${pts[0].x-20}" y1="${pts[0].y}" x2="${pts[0].x+20}" y2="${pts[0].y}" stroke="${col}" stroke-width="1.5" stroke-dasharray="4,3" opacity="0.5"/>`;
    }
    pts.forEach(p=>{
      svg+=`<circle cx="${p.x}" cy="${p.y}" r="4.5" fill="${col}" stroke="#090f1c" stroke-width="1.8"/>`;
      allLabels.push({x:p.x,y:p.y,v:p.v,col,mi:MONTHS.indexOf(p.mo)});
    });
  });
  // anti-overlap label placement: stack per month with min gap
  const byMonth={};
  allLabels.forEach(L=>{(byMonth[L.mi]=byMonth[L.mi]||[]).push(L);});
  Object.values(byMonth).forEach(list=>{
    list.sort((a,b)=>a.y-b.y);
    const MINGAP=17;
    let lastY=-999;
    list.forEach(L=>{
      let ly=L.y-12;
      if(ly<lastY+MINGAP) ly=lastY+MINGAP;
      lastY=ly;
      const lbl=`${L.v}d`;const lw=lbl.length*6.5+12;
      svg+=`<line x1="${L.x}" y1="${L.y}" x2="${L.x}" y2="${ly+3}" stroke="${L.col}" stroke-width="1" opacity="0.35"/>`;
      svg+=`<rect x="${L.x-lw/2}" y="${ly-6}" width="${lw}" height="15" rx="3" fill="#060e1c" opacity="0.96" stroke="${L.col}" stroke-opacity="0.3" stroke-width="0.8"/>`;
      svg+=`<text x="${L.x}" y="${ly+5}" text-anchor="middle" fill="#ffffff" font-size="10.5" font-family="DM Mono" font-weight="700">${lbl}</text>`;
    });
  });
  svg+='</svg>';el.innerHTML=svg;svgTips(el);
  const lg=document.getElementById('lg-'+brand);
  if(lg)lg.innerHTML=keys.map(k=>`<span class="li"><span class="sw" style="background:${META[k]?META[k].color:'#93c5fd'}"></span>${META[k]?META[k].label:k}</span>`).join('');
}

// ── FUNNEL ───────────────────────────────────────────────────
function drawFunnel(brand){
  const d=DATA[brand];const el=document.getElementById('f-'+brand);if(!el||!d)return;
  const a=d.avgs||{};
  const items=[{l:'Invoice→GL',v:a.lt_inv_gl,c:'#93c5fd'},{l:'ARR→ID Reg',v:a.lt_arr_id,c:'#818cf8'},{l:'ID→CC (R/Y)',v:a.lt_id_cc_ry,c:'#f59e0b'},{l:'CC→NF',v:a.lt_cc_nf,c:'#f472b6'},{l:'NF→POD',v:a.lt_nf_pod,c:'#22c55e'}].filter(x=>x.v!=null);
  if(!items.length){el.innerHTML='<div style="color:#2a3f5f;font-size:10px;padding:8px">No data</div>';return;}
  const max=Math.max(...items.map(i=>i.v),1);
  el.innerHTML=items.map(it=>`<div class="frow"><div class="fflbl">${it.l}</div><div class="ffbg"><div class="ffbar" style="width:${it.v/max*100}%;background:${it.c}">${it.v}d</div></div><div class="ffavg">${it.v}d</div></div>`).join('')+`<div style="margin-top:9px;font-size:8.5px;color:#263545;text-align:center;letter-spacing:1px;text-transform:uppercase">Avg — filtered processes</div>`;
}

// ── BAR LIST ─────────────────────────────────────────────────
function drawBlist(id,items){
  const el=document.getElementById(id);if(!el)return;
  if(!items||!items.length){el.innerHTML='<div style="color:#263545;font-size:10px;padding:8px 0">No data</div>';return;}
  const max=Math.max(...items.map(s=>s.v),1);
  el.innerHTML=items.map(s=>`<div class="brow"><div class="bnm" title="${s.n}">${s.n}</div><div class="bbg"><div class="bfill" style="width:${s.v/max*100}%;background:${s.c}"></div></div><div class="bval">${s.v}</div></div>`).join('');
}

// ── VOLUME ───────────────────────────────────────────────────
function drawVol(brand){
  const d=DATA[brand];const el=document.getElementById('vol-'+brand);if(!el||!d)return;
  const vol=d.vol||{};const col=d.color;
  const mos=Object.keys(vol).filter(m=>vol[m]>0).sort();
  if(!mos.length){el.innerHTML='<div style="color:#2a3f5f;font-size:10px;padding:8px">No data</div>';return;}
  const w=el.clientWidth||340,h=el.clientHeight||180;
  const m={t:24,r:12,b:36,l:38};
  const maxV=Math.max(...mos.map(mo=>vol[mo]),1);
  const gap=(w-m.l-m.r)/mos.length;const bw=Math.max(18,gap*0.6);
  let svg=`<svg viewBox="0 0 ${w} ${h}">`;
  for(let i=0;i<=3;i++){const y=h-m.b-(h-m.t-m.b)*i/3;svg+=`<line class="gl" x1="${m.l}" y1="${y}" x2="${w-m.r}" y2="${y}"/><text class="ax" x="${m.l-5}" y="${y+4}" text-anchor="end">${Math.round(maxV*i/3)}</text>`;}
  mos.forEach((mo,i)=>{
    const v=vol[mo];const x=m.l+i*gap+(gap-bw)/2;const bh=(h-m.t-m.b)*v/maxV;const y=h-m.b-bh;
    svg+=`<rect x="${x}" y="${y}" width="${bw}" height="${bh}" rx="3" fill="${col}" fill-opacity="0.78" data-tip="${mo}: <strong>${v.toLocaleString()}</strong>"/>`;
    svg+=`<text class="ax" x="${x+bw/2}" y="${h-3}" text-anchor="middle">${MLABELS[mo]||mo.slice(5)}</text>`;
    const lbl=v.toLocaleString();const lw=lbl.length*7+12;
    svg+=`<rect x="${x+bw/2-lw/2}" y="${y-17}" width="${lw}" height="14" rx="3" fill="#0a1428" opacity="0.9"/>`;
    svg+=`<text x="${x+bw/2}" y="${y-5}" text-anchor="middle" fill="${col}" font-size="10" font-family="DM Mono" font-weight="700">${lbl}</text>`;
  });
  svg+='</svg>';el.innerHTML=svg;svgTips(el);
}

// ── SLA DONUTS ───────────────────────────────────────────────
function drawSla(brand){
  const d=DATA[brand];const el=document.getElementById('sla-'+brand);if(!el||!d)return;
  const sla=d.sla||{};
  const keys=['lt_inv_gl','lt_arr_id','lt_id_cc_ry','lt_cc_nf'];
  el.innerHTML=keys.map(k=>{
    const s=sla[k]||{ok:0,br:0,na:0};const meas=s.ok+s.br;
    const pct=meas>0?Math.round(s.ok/meas*100):0;const p=pc(pct);
    const R=33,cx=39,cy=39,sw=8,circ=2*Math.PI*R;
    const dash=(pct/100)*circ,g2=circ-dash;
    const donut=`<svg viewBox="0 0 78 78" style="transform:rotate(-90deg)"><circle cx="${cx}" cy="${cy}" r="${R}" fill="none" stroke="#080e1c" stroke-width="${sw}"/><circle cx="${cx}" cy="${cy}" r="${R}" fill="none" stroke="${p}" stroke-width="${sw}" stroke-dasharray="${dash.toFixed(1)} ${g2.toFixed(1)}" stroke-linecap="round"/></svg>`;
    const isIdc=k==='lt_id_cc_ry';
    const m2=META[k]||{label:k,sla:1};
    const tip=`<strong>${m2.label}</strong> (SLA ≤${m2.sla}d)${isIdc?'<br><em style="color:#3a5070;font-size:9px">RED/YELLOW channel only</em>':''}<br><span style="color:#93c5fd">Within: ${s.ok}</span><br><span style="color:#60a5fa">Breaching: ${s.br}</span>${s.na>0?`<br><span style="color:#263545">Not reached: ${s.na}</span>`:''}<br>Compliance: <strong style="color:${p}">${pct}%</strong>`;
    return `<div class="sla-card" onmouseenter="showTip(event,this.dataset.tip)" onmousemove="mvTip(event)" onmouseleave="hideTip()" data-tip="${tip.replace(/"/g,'&quot;')}">
      <div class="sttl">${m2.label}</div><div class="stgt">SLA ≤ ${m2.sla}d</div>
      <div class="sdn">${donut}<div class="sctr"><div class="spct" style="color:${p}">${pct}%</div><div class="ssub">in SLA</div></div></div>
      <div class="scts"><span style="color:${p}">✓${s.ok}</span> <span style="color:#60a5fa">✗${s.br}</span>${s.na>0?` <span style="color:#1e3050">–${s.na}</span>`:''}</div>
      <div class="sbar2"><div class="sfill" style="width:${pct}%;background:${p}"></div></div>
      ${isIdc?'<div style="font-size:7px;color:#1e3050;margin-top:3px">RED / YELLOW channel only</div>':''}
    </div>`;
  }).join('');
}

// ── NF→POD ───────────────────────────────────────────────────
function drawNfPod(brand){
  const d=DATA[brand];const el=document.getElementById('nfpod-'+brand);if(!el||!d)return;
  const nfpod=d.nfpod||{SAO:{ok:0,br:0,sla:2},RIO:{ok:0,br:0,sla:3},OTHER:{ok:0,br:0,sla:5}};
  const locs=[{key:'SAO',label:'São Paulo',sla:2,col:'#93c5fd'},{key:'RIO',label:'Rio de Janeiro',sla:3,col:'#60a5fa'},{key:'OTHER',label:'Other States',sla:5,col:'#3b82f6'}];
  el.innerHTML=locs.map(loc=>{
    const s=nfpod[loc.key]||{ok:0,br:0};const meas=s.ok+s.br;
    const pct=meas>0?Math.round(s.ok/meas*100):null;const p=pct!=null?pc(pct):'#1e3050';
    const dp=pct!=null?pct:0;
    const R=33,cx=39,cy=39,sw=8,circ=2*Math.PI*R;
    const dash=(dp/100)*circ,g2=circ-dash;
    const donut=`<svg viewBox="0 0 78 78" style="transform:rotate(-90deg)"><circle cx="${cx}" cy="${cy}" r="${R}" fill="none" stroke="#080e1c" stroke-width="${sw}"/><circle cx="${cx}" cy="${cy}" r="${R}" fill="none" stroke="${p}" stroke-width="${sw}" stroke-dasharray="${dash.toFixed(1)} ${g2.toFixed(1)}" stroke-linecap="round"/></svg>`;
    const tip=`<strong>${loc.label}</strong><br>SLA: ≤${loc.sla}d<br><span style="color:#93c5fd">Within: ${s.ok}</span><br><span style="color:#60a5fa">Breaching: ${s.br}</span><br>Compliance: <strong style="color:${p}">${pct!=null?pct+'%':'No data'}</strong>`;
    return `<div class="sla-card" onmouseenter="showTip(event,this.dataset.tip)" onmousemove="mvTip(event)" onmouseleave="hideTip()" data-tip="${tip.replace(/"/g,'&quot;')}">
      <div class="sttl">${loc.label}</div><div class="stgt">SLA ≤ ${loc.sla}d</div>
      <div class="sdn">${donut}<div class="sctr"><div class="spct" style="color:${p}">${pct!=null?pct+'%':'—'}</div><div class="ssub">in SLA</div></div></div>
      <div class="scts"><span style="color:${p}">✓${s.ok}</span> <span style="color:#60a5fa">✗${s.br}</span></div>
      <div class="sbar2"><div class="sfill" style="width:${dp}%;background:${p}"></div></div>
    </div>`;
  }).join('');
}

// ── MODAL ────────────────────────────────────────────────────
const BNAMES={balenciaga:'Balenciaga',bottega:'Bottega Veneta',ysl:'Saint Laurent',gucci:'Gucci',total:'Total (All Brands)'};
function openMod(brand){
  const d=DATA[brand];if(!d)return;
  const totInv=d.del_list.reduce((s,x)=>s+x.invoices,0);
  document.getElementById('mtitle').innerHTML=`<span>${d.del_list.length} Delayed Shipment${d.del_list.length!==1?'s':''}</span> — ${BNAMES[brand]}`;
  document.getElementById('msub').textContent=`${totInv} invoice${totInv!==1?'s':''} · Lead times vs SLA · White line = SLA threshold`;
  document.getElementById('mbody').innerHTML=d.del_list.map(ship=>{
    const breached=ship.milestones.filter(m=>m.val!=null&&m.val>m.sla);
    const maxV=Math.max(...ship.milestones.filter(m=>m.val!=null).map(m=>Math.max(m.val,m.sla)),1)*1.15;
    const msRows=ship.milestones.map(m=>{
      if(m.val==null)return`<div class="msrow"><div class="mslbl">${m.key}</div><div class="mstrack" style="background:#060c18"><span style="position:absolute;left:8px;top:50%;transform:translateY(-50%);font-size:8.5px;color:#2a3f5f;font-style:italic">No data</span></div><div class="msval" style="color:#2a3f5f">—</div><div class="msslal">≤${m.sla}d</div></div>`;
      const fp=Math.min(m.val/maxV*100,100);const sp=Math.min(m.sla/maxV*100,100);
      return`<div class="msrow"><div class="mslbl">${m.key}</div><div class="mstrack"><div class="msfill" style="width:${fp}%;background:${m.color}">${m.val>=4?m.val+'d':''}</div><div class="msline" style="left:${sp}%"></div></div><div class="msval" style="color:${m.color}">${m.val}d</div><div class="msslal">≤${m.sla}d</div></div>`;
    }).join('');
    return`<div class="scard"><div class="shdr"><div class="sid">${ship.shipment}</div><div class="smeta"><span class="stag loc">${ship.location}</span><span class="stag inv">${ship.invoices} invoice${ship.invoices!==1?'s':''}</span><span class="stag br">${breached.length} milestone${breached.length!==1?'s':''} breached</span></div></div><div style="font-size:8px;color:#2a3f5f;letter-spacing:1.5px;text-transform:uppercase;font-weight:700;margin-bottom:8px;display:flex;align-items:center;gap:8px">Milestone Performance<span style="flex:1;height:1px;background:#0f1e30"></span><span style="color:#1e3050;font-style:italic;text-transform:none;letter-spacing:0">White line = SLA target</span></div><div class="mschart">${msRows}</div></div>`;
  }).join('');
  document.getElementById('moverlay').classList.add('open');
  document.body.style.overflow='hidden';
}
function closeMod(){document.getElementById('moverlay').classList.remove('open');document.body.style.overflow='';}
document.addEventListener('keydown',e=>{if(e.key==='Escape')closeMod();});

// ── PILLS ────────────────────────────────────────────────────
['balenciaga','bottega','ysl','gucci','total'].forEach(brand=>{
  const w=document.getElementById('pills-'+brand);if(!w)return;
  w.addEventListener('click',e=>{
    const pill=e.target.closest('.pill');if(!pill)return;
    tModes[brand]=pill.dataset.mode;
    w.querySelectorAll('.pill').forEach(p=>p.classList.remove('active'));
    pill.classList.add('active');
    drawTrend(brand);
  });
});

// ── RENDER ALL ───────────────────────────────────────────────
function renderAll(){
  ['balenciaga','bottega','ysl','gucci','total'].forEach(brand=>{
    const d=DATA[brand];if(!d)return;
    const d_sts = DATA_ALL_STS[brand]; if(!d_sts)return;
    drawSbar(brand);drawKpis(brand);drawPipe(brand);
    drawTrend(brand);drawFunnel(brand);
    drawBlist('stage-'+brand,d.stages);
    drawBlist('status-'+brand,d_sts.statuses);
    //drawBlist('status-'+brand,d.statuses);
    drawVol(brand);drawSla(brand);drawNfPod(brand);
  });
}
window.addEventListener('resize',renderAll);
recompute();
renderAll();
"""


def montar_html(rows, delivered_rows, status_step_rows, timestamp_str):
    """Monta o arquivo HTML final injetando os dados no template."""
    rows_json = json.dumps(rows, separators=(",", ":"), ensure_ascii=False)
    delivered_rows_json = json.dumps(delivered_rows, separators=(",", ":"), ensure_ascii=False)
    status_step_rows_json = json.dumps(status_step_rows, separators=(",", ":"), ensure_ascii=False)

    # ── intervalo real de datas presente nos dados ──────────────────────
    datas = computar_datas(rows)
    months, mlabels = computar_months(datas)
    week_ranges, week_options = computar_weeks(datas)

    if datas:
        dmin, dmax = min(datas), max(datas)

        def fmt(d):
            return f"{MESES_ABREV[f'{d.month:02d}']} {d.day}, {d.year}"

        data_range_label = f"{fmt(dmin)} — {fmt(dmax)}"
    else:
        data_range_label = "—"

    last_update_label = f"{timestamp_str} · Brasília (BR)"

    week_options_html = "\n      ".join(
        f'<option value="{chave}">{rotulo}</option>' for chave, rotulo in week_options
    )

    months_json = json.dumps(months, separators=(",", ":"))
    mlabels_json = json.dumps(mlabels, separators=(",", ":"), ensure_ascii=False)
    week_ranges_json = json.dumps(week_ranges, separators=(",", ":"))

    body = (
        HTML_BODY_TEMPLATE
        .replace("{DATA_RANGE_LABEL}", data_range_label)
        .replace("{LAST_UPDATE_LABEL}", last_update_label)
        .replace("{WEEK_OPTIONS_HTML}", week_options_html)
    )

    js_after_data = (
        JS_AFTER_DATA_TEMPLATE
        .replace("{WEEK_RANGES_JSON}", week_ranges_json)
        .replace("{MONTHS_JSON}", months_json)
        .replace("{MLABELS_JSON}", mlabels_json)
    )

    data_block = f"""// ── DADOS — atualizado automaticamente pelo script Python ──────
// Gerado em: {timestamp_str} (fuso Brasilia - BR)
// Campos por linha (16 indices):
// [0]brand [1]ship [2]boxes [3]items [4]status [5]loc [6]channel
// [7]status_lead [8]inv_receipt(YYYY-MM-DD) [9]eta(YYYY-MM-DD|null)
// [10]lt_ig [11]lt_ai [12]lt_ic [13]lt_cn [14]lt_np [15]invoice
const ALL_ROWS={rows_json};
// Linhas com status == DELIVERED (mesmo formato de ALL_ROWS), usadas
// apenas para os KPIs adicionais de "ja entregue" no Overview:
const DELIVERED_ROWS={delivered_rows_json};
// Linhas com os outros status 'WAITING ARRIVAL', 'WAITING GL' e 'WAITING PRE ALERT'  (mesmo formato de ALL_ROWS),
// usadas na etapa 'STATUS':
const ALL_STS_ROWS={status_step_rows_json};
// ── FIM DOS DADOS ────────────────────────────────────────────"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Kering Import Operations | All Open Processes</title>
<style>
{CSS_TEMPLATE}
</style>
</head>
<body>
{body}
<script>
{data_block}
{js_after_data}
</script>
</body>
</html>"""
    return html


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Gera o dashboard Kering Import Operations a partir da planilha unificada."
    )
    parser.add_argument("planilha", help="Caminho para o arquivo .xlsx da base unificada")
    parser.add_argument(
        "--out", "-o",
        default="import_operations_all_open.html",
        help="Nome do arquivo HTML de saida (padrao: import_operations_all_open.html)",
    )
    args = parser.parse_args()

    caminho = Path(args.planilha)
    if not caminho.exists():
        print(f"ERRO: arquivo nao encontrado: {caminho}")
        sys.exit(1)

    rows = extrair_rows(caminho)
    if not rows:
        print("ERRO: nenhuma linha valida foi extraida. Verifique o arquivo e os filtros.")
        sys.exit(1)

    imprimir_resumo(rows)

    delivered_rows = extrair_rows(caminho, status_filtro=STATUS_DELIVERED, silencioso=True)

    status_step_rows = extrair_rows(caminho, status_filtro=STATUS_VALIDOS_STS, silencioso=True)

    print(f"\n  Linhas DELIVERED encontradas: {len(delivered_rows)}")

    timestamp_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%Mh")
    html = montar_html(rows, delivered_rows, status_step_rows, timestamp_str)

    out_path = Path(args.out)
    out_path.write_text(html, encoding="utf-8")

    tamanho_kb = out_path.stat().st_size // 1024
    print(f"\nDashboard gerado: {out_path} ({tamanho_kb} kb)")
    print("Abra este arquivo .html diretamente no navegador.")


if __name__ == "__main__":
    main()
